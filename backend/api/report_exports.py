"""Utilities to export accounting reports as Excel workbooks or PDFs."""

from collections import Counter, defaultdict
from decimal import Decimal
from io import BytesIO
from typing import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


__all__ = [
    "generate_sales_report_workbook",
    "generate_sales_report_pdf",
    "generate_profit_loss_workbook",
    "generate_profit_loss_pdf",
    "generate_customer_balance_workbook",
    "generate_customer_balance_pdf",
]


HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CURRENCY_NUMBER_FORMAT = "#,##0.00"


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _auto_size_columns(worksheet) -> None:
    """Adjust column widths to fit their content nicely."""

    for column_cells in worksheet.columns:
        try:
            column_letter = get_column_letter(column_cells[0].column)
        except AttributeError:  # pragma: no cover - defensive for older openpyxl
            column_letter = get_column_letter(column_cells[0].column_letter)

        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def _format_currency(value: Decimal) -> str:
    return f"{_to_decimal(value):,.2f}"


STATUS_LABELS = {
    "owes_us": "Customers Owing Us",
    "we_owe_them": "Customers We Owe",
    "settled": "Settled Customers",
}

STATUS_ORDER = ("owes_us", "we_owe_them", "settled")


def _summarize_customer_balances(customers: Sequence[Mapping]) -> tuple[Counter, dict[str, dict[str, Decimal]]]:
    """Return status counts and currency totals for the customer balances."""

    status_counts: Counter = Counter()
    currency_totals: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"owe_us": Decimal("0"), "we_owe_them": Decimal("0")}
    )

    for customer in customers:
        status = (customer.get("status") or "settled").lower()
        status_counts[status] += 1

        balance = _to_decimal(customer.get("balance"))
        currency = customer.get("currency") or "USD"

        if balance > 0:
            currency_totals[currency]["owe_us"] += balance
        elif balance < 0:
            currency_totals[currency]["we_owe_them"] += abs(balance)

    return status_counts, currency_totals


def generate_sales_report_workbook(sales: Sequence, start_date: str, end_date: str) -> bytes:
    """Return an Excel workbook representing the sales report."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales Report"

    worksheet["A1"] = "Sales Report"
    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet["A2"] = f"Period: {start_date} to {end_date}"
    worksheet["A2"].font = Font(italic=True)
    worksheet.append([])

    header = ["#", "Date", "Invoice", "Customer/Supplier", "Currency", "Total Amount"]
    worksheet.append(header)
    for cell in worksheet[worksheet.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    running_total = Decimal("0")

    for index, sale in enumerate(sales, start=1):
        counterparty = ""
        if getattr(sale, "customer", None):
            counterparty = getattr(sale.customer, "name", "")
        elif getattr(sale, "supplier", None):
            counterparty = getattr(sale.supplier, "name", "")

        amount = _to_decimal(getattr(sale, "total_amount", 0))
        running_total += amount

        worksheet.append([
            index,
            sale.sale_date.strftime("%Y-%m-%d"),
            sale.invoice_number or "",
            counterparty,
            getattr(sale, "original_currency", ""),
            float(amount),
        ])

        row = worksheet[worksheet.max_row]
        row[0].alignment = Alignment(horizontal="center")
        row[1].alignment = Alignment(horizontal="center")
        row[5].number_format = CURRENCY_NUMBER_FORMAT
        row[5].alignment = Alignment(horizontal="right")

    worksheet.append([])
    worksheet.append(["", "", "", "Grand Total", "", float(running_total)])
    total_row = worksheet[worksheet.max_row]
    total_row[3].font = Font(bold=True)
    total_row[3].fill = TOTAL_FILL
    total_row[5].font = Font(bold=True)
    total_row[5].fill = TOTAL_FILL
    total_row[5].number_format = CURRENCY_NUMBER_FORMAT
    total_row[5].alignment = Alignment(horizontal="right")

    _auto_size_columns(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generate_sales_report_pdf(sales: Sequence, start_date: str, end_date: str) -> bytes:
    """Return a PDF document representing the sales report."""

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="Sales Report",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="RightAlign", alignment=TA_RIGHT))

    story = [
        Paragraph("Sales Report", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(f"Period: {start_date} to {end_date}", styles["Normal"]),
        Spacer(1, 6 * mm),
    ]

    table_data: list[list[str]] = [
        ["#", "Date", "Invoice", "Customer/Supplier", "Currency", "Total Amount"],
    ]

    running_total = Decimal("0")

    for index, sale in enumerate(sales, start=1):
        counterparty = ""
        if getattr(sale, "customer", None):
            counterparty = getattr(sale.customer, "name", "")
        elif getattr(sale, "supplier", None):
            counterparty = getattr(sale.supplier, "name", "")

        amount = _to_decimal(getattr(sale, "total_amount", 0))
        running_total += amount

        table_data.append([
            str(index),
            sale.sale_date.strftime("%Y-%m-%d"),
            sale.invoice_number or "",
            counterparty,
            getattr(sale, "original_currency", ""),
            _format_currency(amount),
        ])

    table_data.append(["", "", "", "Grand Total", "", _format_currency(running_total)])

    table = Table(
        table_data,
        colWidths=[15 * mm, 30 * mm, 35 * mm, 90 * mm, 25 * mm, 30 * mm],
        repeatRows=1,
    )
    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -2), "CENTER"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -2), "LEFT"),
        ("FONTNAME", (3, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F2F2F2")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
    ])
    table.setStyle(table_style)
    story.append(table)

    document.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def generate_profit_loss_workbook(report_data: dict) -> bytes:
    """Return an Excel workbook for the profit and loss statement."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Profit & Loss"

    start_date = report_data.get("start_date", "")
    end_date = report_data.get("end_date", "")

    worksheet["A1"] = "Profit & Loss Statement"
    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet["A2"] = f"Period: {start_date} to {end_date}".strip()
    worksheet["A2"].font = Font(italic=True)
    worksheet.append([])

    header = ["Section", "Category", "Amount"]
    worksheet.append(header)
    for cell in worksheet[worksheet.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.append([
        "Revenue",
        "Total Revenue",
        float(_to_decimal(report_data.get("total_revenue"))),
    ])

    expenses = report_data.get("expenses_breakdown", []) or []
    for entry in expenses:
        worksheet.append([
            "Expenses",
            entry.get("category__name") or "Uncategorized",
            float(_to_decimal(entry.get("total"))),
        ])

    worksheet.append([
        "Expenses",
        "Total Expenses",
        float(_to_decimal(report_data.get("total_expenses"))),
    ])
    worksheet.append([
        "Summary",
        "Net Profit",
        float(_to_decimal(report_data.get("net_profit"))),
    ])

    for row in worksheet.iter_rows(min_row=5, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = CURRENCY_NUMBER_FORMAT
            cell.alignment = Alignment(horizontal="right")

    last_row = worksheet.max_row
    worksheet[f"A{last_row}"].font = Font(bold=True)
    worksheet[f"B{last_row}"].font = Font(bold=True)
    worksheet[f"C{last_row}"].font = Font(bold=True)
    worksheet[f"A{last_row}"].fill = TOTAL_FILL
    worksheet[f"B{last_row}"].fill = TOTAL_FILL
    worksheet[f"C{last_row}"].fill = TOTAL_FILL

    _auto_size_columns(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generate_profit_loss_pdf(report_data: dict) -> bytes:
    """Return a PDF version of the profit and loss statement."""

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        title="Profit & Loss Statement",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="AmountRight", alignment=TA_RIGHT))

    start_date = report_data.get("start_date", "")
    end_date = report_data.get("end_date", "")

    story = [
        Paragraph("Profit & Loss Statement", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(f"Period: {start_date} to {end_date}".strip(), styles["Normal"]),
        Spacer(1, 6 * mm),
    ]

    table_data: list[list[str]] = [["Section", "Category", "Amount"]]

    table_data.append([
        "Revenue",
        "Total Revenue",
        _format_currency(report_data.get("total_revenue")),
    ])

    expenses = report_data.get("expenses_breakdown", []) or []
    for entry in expenses:
        table_data.append([
            "Expenses",
            entry.get("category__name") or "Uncategorized",
            _format_currency(entry.get("total")),
        ])

    table_data.append([
        "Expenses",
        "Total Expenses",
        _format_currency(report_data.get("total_expenses")),
    ])
    table_data.append([
        "Summary",
        "Net Profit",
        _format_currency(report_data.get("net_profit")),
    ])

    table = Table(table_data, colWidths=[40 * mm, 90 * mm, 35 * mm], repeatRows=1)
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
    ]

    # Highlight specific rows for readability
    revenue_row = 1
    table_style.append(("BACKGROUND", (0, revenue_row), (-1, revenue_row), colors.HexColor("#E2F0D9")))

    total_expenses_row = revenue_row + len(expenses) + 1
    if expenses:
        expenses_start = revenue_row + 1
        expenses_end = expenses_start + len(expenses) - 1
        table_style.append(("BACKGROUND", (0, expenses_start), (-1, expenses_end), colors.HexColor("#FCE4D6")))

    table_style.append(("BACKGROUND", (0, total_expenses_row), (-1, total_expenses_row), colors.HexColor("#F8CBAD")))

    net_profit_row = total_expenses_row + 1
    table_style.append(("BACKGROUND", (0, net_profit_row), (-1, net_profit_row), colors.HexColor("#D9E1F2")))
    table_style.append(("FONTNAME", (0, net_profit_row), (-1, net_profit_row), "Helvetica-Bold"))

    table.setStyle(TableStyle(table_style))
    story.append(table)

    document.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def generate_customer_balance_workbook(customers: Sequence[Mapping]) -> bytes:
    """Return an Excel workbook summarising customer balances."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Customer Balances"

    worksheet["A1"] = "Customer Balance Report"
    worksheet["A1"].font = Font(size=14, bold=True)

    status_counts, currency_totals = _summarize_customer_balances(customers)

    worksheet.append([])
    worksheet.append(["Status Summary"])
    worksheet[worksheet.max_row][0].font = Font(bold=True)
    worksheet.append(["Status", "Customers"])
    header_row = worksheet[worksheet.max_row]
    for cell in header_row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for key in STATUS_ORDER:
        worksheet.append([STATUS_LABELS[key], int(status_counts.get(key, 0))])

    worksheet.append([])
    worksheet.append(["Currency Totals"])
    worksheet[worksheet.max_row][0].font = Font(bold=True)
    worksheet.append(["Currency", "Customers Owing Us", "We Owe Them"])
    currency_header_row = worksheet[worksheet.max_row]
    for cell in currency_header_row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    currency_data_start = worksheet.max_row + 1
    if currency_totals:
        for currency in sorted(currency_totals):
            totals = currency_totals[currency]
            worksheet.append([
                currency,
                float(totals["owe_us"]),
                float(totals["we_owe_them"]),
            ])
    else:
        worksheet.append(["—", 0.0, 0.0])

    currency_data_end = worksheet.max_row
    for row in worksheet.iter_rows(
        min_row=currency_data_start, max_row=currency_data_end, min_col=2, max_col=3
    ):
        for cell in row:
            cell.number_format = CURRENCY_NUMBER_FORMAT
            cell.alignment = Alignment(horizontal="right")

    worksheet.append([])
    worksheet.append(["Customer Details"])
    worksheet[worksheet.max_row][0].font = Font(bold=True)
    worksheet.append(["Name", "Email", "Phone", "Currency", "Balance", "Status"])
    details_header_row = worksheet[worksheet.max_row]
    for cell in details_header_row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if customers:
        for customer in customers:
            balance_decimal = _to_decimal(customer.get("balance"))
            worksheet.append(
                [
                    customer.get("name") or "",
                    customer.get("email") or "",
                    customer.get("phone") or "",
                    customer.get("currency") or "",
                    float(balance_decimal),
                    STATUS_LABELS.get((customer.get("status") or "").lower(), ""),
                ]
            )
            row = worksheet[worksheet.max_row]
            row[4].number_format = CURRENCY_NUMBER_FORMAT
            row[4].alignment = Alignment(horizontal="right")
    else:
        worksheet.append(["No customer balances available.", "", "", "", "", ""])

    _auto_size_columns(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generate_customer_balance_pdf(customers: Sequence[Mapping]) -> bytes:
    """Return a PDF document summarising customer balances."""

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        title="Customer Balance Report",
    )

    styles = getSampleStyleSheet()

    story = [
        Paragraph("Customer Balance Report", styles["Title"]),
        Spacer(1, 6 * mm),
    ]

    status_counts, currency_totals = _summarize_customer_balances(customers)

    status_data: list[list[str]] = [["Status", "Customers"]]
    for key in STATUS_ORDER:
        status_data.append([STATUS_LABELS[key], str(int(status_counts.get(key, 0)))])

    status_table = Table(status_data, colWidths=[80 * mm, 30 * mm])
    status_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 1), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
            ]
        )
    )
    story.append(status_table)
    story.append(Spacer(1, 4 * mm))

    currency_data: list[list[str]] = [["Currency", "Customers Owing Us", "We Owe Them"]]
    if currency_totals:
        for currency in sorted(currency_totals):
            totals = currency_totals[currency]
            currency_data.append(
                [
                    currency,
                    _format_currency(totals["owe_us"]),
                    _format_currency(totals["we_owe_them"]),
                ]
            )
    else:
        currency_data.append(["—", _format_currency(Decimal("0")), _format_currency(Decimal("0"))])

    currency_table = Table(currency_data, colWidths=[40 * mm, 55 * mm, 55 * mm])
    currency_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 1), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
            ]
        )
    )
    story.append(currency_table)
    story.append(Spacer(1, 6 * mm))

    details_data: list[list[str]] = [["Name", "Email", "Phone", "Currency", "Balance", "Status"]]
    if customers:
        for customer in customers:
            details_data.append(
                [
                    customer.get("name") or "",
                    customer.get("email") or "",
                    customer.get("phone") or "",
                    customer.get("currency") or "",
                    _format_currency(customer.get("balance")),
                    STATUS_LABELS.get((customer.get("status") or "").lower(), ""),
                ]
            )
    else:
        details_data.append(["No customer balances available.", "", "", "", "", ""])

    details_table = Table(
        details_data,
        colWidths=[40 * mm, 55 * mm, 35 * mm, 25 * mm, 30 * mm, 50 * mm],
        repeatRows=1,
    )
    details_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 0), (-1, 0), 4),
                ("TOPPADDING", (0, 1), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
            ]
        )
    )
    story.append(details_table)

    document.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf
