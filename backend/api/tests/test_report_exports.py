"""Tests covering PDF and Excel exports for accounting reports."""

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from ..models import (
    Customer,
    Expense,
    ExpenseCategory,
    Product,
    Sale,
    Warehouse,
    WarehouseInventory,
)


class ReportExportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="reportuser", password="pw")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

        self.customer = Customer.objects.create(name="Acme Corp", created_by=self.user)
        self.sale = Sale.objects.create(
            customer=self.customer,
            created_by=self.user,
            invoice_number="INV-001",
            original_amount=Decimal("100.00"),
            exchange_rate=Decimal("1"),
        )
        self.sale.refresh_from_db()
        self.sale_date = self.sale.sale_date
        self.report_end_date = (self.sale_date + timedelta(days=1)).strftime("%Y-%m-%d")

        Customer.objects.create(
            name="Globex", created_by=self.user, currency="EUR", open_balance=Decimal("-25.00")
        )
        Customer.objects.create(
            name="Initech", created_by=self.user, currency="USD", open_balance=Decimal("0.00")
        )

        category = ExpenseCategory.objects.create(name="Office", created_by=self.user)
        self.expense = Expense.objects.create(
            category=category,
            amount=Decimal("40.00"),
            original_amount=Decimal("40.00"),
            expense_date=date(2024, 1, 2),
            created_by=self.user,
        )
        self.expense.refresh_from_db()

        self.product = Product.objects.create(
            name="Widget",
            description="Test product",
            sku="W-001",
            purchase_price=Decimal("12.50"),
            sale_price=Decimal("20.00"),
            created_by=self.user,
        )

        self.default_warehouse = Warehouse.get_default(self.user)
        WarehouseInventory.adjust_stock(self.product, self.default_warehouse, Decimal("5"))

        self.secondary_warehouse = Warehouse.objects.create(
            name="Secondary Warehouse",
            created_by=self.user,
        )
        WarehouseInventory.adjust_stock(self.product, self.secondary_warehouse, Decimal("3"))

    def _assert_pdf_response(self, response):
        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_sales_report_excel_contains_data(self):
        response = self.client.get(
            "/api/reports/sales/",
            {
                "start_date": "2023-01-01",
                "end_date": self.report_end_date,
                "export_format": "xlsx",
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet["B5"].value, self.sale_date.strftime("%Y-%m-%d"))
        self.assertEqual(worksheet["C5"].value, self.sale.invoice_number)
        self.assertEqual(worksheet["D5"].value, self.customer.name)
        self.assertAlmostEqual(worksheet["F5"].value, float(self.sale.total_amount))
        self.assertAlmostEqual(worksheet["F7"].value, float(self.sale.total_amount))

    def test_sales_report_pdf_download(self):
        response = self.client.get(
            "/api/reports/sales/",
            {
                "start_date": "2023-01-01",
                "end_date": self.report_end_date,
                "export_format": "pdf",
            },
        )

        self._assert_pdf_response(response)

    def test_profit_loss_excel_contains_totals(self):
        self.assertEqual(self.expense.amount, Decimal("40.00"))
        json_response = self.client.get(
            "/api/reports/profit-loss/",
            {
                "start_date": "2023-01-01",
                "end_date": self.report_end_date,
            },
        )
        payload = json_response.json()
        self.assertAlmostEqual(float(payload["total_expenses"]), 40.0)
        self.assertAlmostEqual(float(payload["net_profit"]), float(self.sale.total_amount) - 40.0)

        response = self.client.get(
            "/api/reports/profit-loss/",
            {
                "start_date": "2023-01-01",
                "end_date": self.report_end_date,
                "export_format": "xlsx",
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet["A5"].value, "Revenue")
        self.assertAlmostEqual(worksheet["C5"].value, float(self.sale.total_amount))
        rows = list(worksheet.iter_rows(values_only=True))
        total_expenses_row = next(
            (row for row in rows if row and len(row) > 2 and row[1] == "Total Expenses"),
            None,
        )
        self.assertIsNotNone(total_expenses_row)
        self.assertAlmostEqual(total_expenses_row[2], 40.0)

        summary_row = next(
            (row for row in rows if row and row[0] == "Summary"),
            None,
        )
        self.assertIsNotNone(summary_row)
        self.assertAlmostEqual(summary_row[2], float(self.sale.total_amount) - 40.0)

    def test_profit_loss_pdf_download(self):
        response = self.client.get(
            "/api/reports/profit-loss/",
            {
                "start_date": "2023-01-01",
                "end_date": self.report_end_date,
                "export_format": "pdf",
            },
        )

        self._assert_pdf_response(response)

    def test_customer_balance_excel_contains_summary_and_details(self):
        response = self.client.get(
            "/api/reports/customer-balances/",
            {
                "export_format": "xlsx",
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet["A1"].value, "Customer Balance Report")

        rows = list(worksheet.iter_rows(values_only=True))
        status_counts = {
            row[0]: row[1]
            for row in rows
            if row and row[0] in {"Customers Owing Us", "Customers We Owe", "Settled Customers"}
        }
        self.assertEqual(status_counts.get("Customers Owing Us"), 1)
        self.assertEqual(status_counts.get("Customers We Owe"), 1)
        self.assertEqual(status_counts.get("Settled Customers"), 1)

        currency_rows = {
            row[0]: row
            for row in rows
            if row and row[0] in {"USD", "EUR"}
        }
        self.assertAlmostEqual(currency_rows["USD"][1], float(self.sale.total_amount))
        self.assertAlmostEqual(currency_rows["EUR"][2], 25.0)

        customer_rows = [row for row in rows if row and row[0] == self.customer.name]
        self.assertTrue(customer_rows)
        self.assertAlmostEqual(customer_rows[0][4], float(self.sale.total_amount))
        self.assertEqual(customer_rows[0][5], "Customers Owing Us")

    def test_customer_balance_pdf_download(self):
        response = self.client.get(
            "/api/reports/customer-balances/",
            {
                "export_format": "pdf",
            },
        )

        self._assert_pdf_response(response)

    def test_inventory_report_json_includes_totals(self):
        response = self.client.get("/api/reports/inventory/")

        self.assertEqual(response.status_code, 200, response.content)
        data = response.json()
        product_row = next((item for item in data if item["id"] == self.product.id), None)
        self.assertIsNotNone(product_row, data)
        self.assertEqual(product_row["name"], self.product.name)
        self.assertEqual(product_row["sku"], self.product.sku)
        self.assertEqual(Decimal(str(product_row["stock_quantity"])), Decimal("8"))

    def test_inventory_report_excel_contains_totals(self):
        response = self.client.get(
            "/api/reports/inventory/",
            {
                "export_format": "xlsx",
            },
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet["A1"].value, "Inventory Report")
        self.assertEqual(worksheet["B5"].value, self.product.name)
        self.assertAlmostEqual(worksheet["E5"].value, 8.0)
        self.assertAlmostEqual(worksheet["F5"].value, float(self.product.purchase_price))
        self.assertAlmostEqual(worksheet["G5"].value, float(self.product.sale_price))
        self.assertEqual(worksheet["D7"].value, "Totals")
        self.assertAlmostEqual(worksheet["E7"].value, 8.0)

    def test_inventory_report_pdf_download(self):
        response = self.client.get(
            "/api/reports/inventory/",
            {
                "export_format": "pdf",
            },
        )

        self._assert_pdf_response(response)
